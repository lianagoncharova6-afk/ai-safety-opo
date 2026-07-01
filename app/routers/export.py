import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from app.database import get_db
from app import models

router = APIRouter()


def _build_permit_export(db: Session) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Наряды-допуски"
    ws1.append(
        [
            "ID",
            "Номер",
            "Тип",
            "Статус",
            "Объект ОКПО",
            "Описание работ",
            "Ответственный",
            "Руководитель",
            "Дата начала",
            "Дата окончания",
            "Создан",
            "Закрыт",
        ]
    )
    for p in db.query(models.WorkPermit).all():
        ws1.append(
            [
                p.id,
                p.permit_number,
                p.permit_type,
                p.status,
                p.object_okpo.name if p.object_okpo else "",
                p.work_description,
                p.responsible_person.full_name if p.responsible_person else "",
                p.work_manager.full_name if p.work_manager else "",
                str(p.work_start_date),
                str(p.work_end_date),
                str(p.created_at),
                str(p.closed_at) if p.closed_at else "",
            ]
        )

    ws2 = wb.create_sheet("Нарушения")
    ws2.append(
        [
            "ID",
            "Наряд ID",
            "Тип нарушения",
            "Тяжесть",
            "Вероятность",
            "Уровень риска",
            "Обнаружено AI",
            "Камера",
            "Статус",
            "Дата обнаружения",
        ]
    )
    for v in db.query(models.Violation).all():
        ws2.append(
            [
                v.id,
                v.permit_id,
                v.violation_type,
                v.severity,
                v.probability,
                v.risk_level,
                v.detected_by_ai,
                v.camera_id or "",
                v.status,
                str(v.detected_at),
            ]
        )

    ws3 = wb.create_sheet("Матрица рисков")
    ws3.append(
        [
            "ID",
            "Наряд ID",
            "Описание опасности",
            "Вероятность",
            "Тяжесть",
            "Оценка риска",
            "Категория",
            "Меры",
        ]
    )
    for r in db.query(models.RiskMatrix).all():
        ws3.append(
            [
                r.id,
                r.permit_id,
                r.hazard_description,
                r.probability,
                r.severity,
                r.risk_score,
                r.risk_category,
                r.mitigation_measures or "",
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@router.get("/permits-xlsx")
def export_permits_xlsx(db: Session = Depends(get_db)):
    data = _build_permit_export(db)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=naryady_dopuski.xlsx"},
    )
